# University of Delaware: SLAM LAB September 2021
# Created by Dea Harjianto
# Purpose: Compile and calculate all data in combiend files and place into appropriate order files; ease of use for moving data

import os
import sys
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Desc: retrieves file path to access all files in directory
# Params: none
# Returns: filePath (str)
# Side Effects: none
def get_application_path():
    print('Retrieving file path...')
    filePath = ' '
    filePath = sys.executable # get path of executable
    last_dir = filePath.rfind("/") # find the last directory of executable
    filePath = filePath[:last_dir] # index to this last directory
    return filePath


# Desc: Moves necessary data in AVGS ACROSS CODERS in combined file to order file
# Params: filePath (str)
# Returns: none
# Side Effects: reads in participants; open and saves file
def combined_to_order(filePath: str):
    with open(filePath + '/Input/Participants.txt') as p:
        rawLines = p.readlines()
        pcps = [] # To store column names

        i = 1
        temp = []
        for line in rawLines:
            line = line.strip() # remove leading/trailing white spaces
            if line != '':
                temp.append(line)
            else:
                pcps.append(temp)
                temp = []
        pcps.append(temp)

    for pcp in pcps:
        ID = pcp[0]
        print('Moving and calculating data for ' + ID + '...')
        wbCombined = openpyxl.load_workbook(filePath + "/Combined/" + ID + "_FaceTalk_Combined.xlsx")
        wsCombined = wbCombined.worksheets[2]
        if len(pcp) == 3:
            ordnum = pcp[1]
        else:
            ordnum = pcp[4] 
        lang = pcp[len(pcp) - 1]

        orderFile = "Order " + ordnum + "_FaceTalk_" + lang + ".xlsx"
        wbOrder = openpyxl.load_workbook(filePath + "/Order Template/" + orderFile)
        wsOrder = wbOrder.active

        combinedRows = wsCombined.max_row

        # insert participant ID into file
        wsOrder.cell(1, 2).value = ID

        # copying left look
        for i in range(2, combinedRows + 1):
            c = wsCombined.cell(i, 5)
            wsOrder.cell(i + 3, 4).value = c.value

        # copying right look
        for i in range(2, combinedRows + 1):
            c = wsCombined.cell(i,  7)
            wsOrder.cell(i + 3,  3).value = c.value
        
        # copying center look
        for i in range(2, combinedRows +1):
            c = wsCombined.cell(i,  9)
            wsOrder.cell(i + 3,  2).value = c.value

        # begin the calculations!
        prop_calculations(wbOrder)
        wbOrder.save(str(filePath + "/Order/" + ID + "_" + orderFile))

        order_calculations(filePath, ordnum, wbOrder, ID, orderFile)


# Desc: Uses raw looking data to calculate its proportions
# NOTE: trials whose COMBINED looks (L + R) times is LESS THAN 15 are IGNORED (marked red in Excel)
# Params: wbOrder (Workbook)
# Returns: none
# Side Effects: none
def prop_calculations(wbOrder: Workbook):
    redFill = PatternFill(start_color = 'FF0000',
                            end_color = 'FF0000',
                            fill_type = 'solid')
    wsOrder = wbOrder.worksheets[0]
    rCount = 0
    for row in wsOrder.iter_rows():
        rCount = rCount + 1
        if rCount < 5:
            continue
        else:
            center = wsOrder.cell(rCount,  2).value
            left = wsOrder.cell(rCount,  3).value
            right = wsOrder.cell(rCount,  4).value
            totalLook = left + right

            if center != 0:
                wsOrder.cell(rCount,  7).value = center
            elif totalLook > 15:
                sum = left + right
                wsOrder.cell(rCount,  5).value = left/sum       
                wsOrder.cell(rCount,  6).value = right/sum 
            else:
                # ignore all trials whose L + R < 15
                wsOrder.cell(rCount,  1).fill = redFill
                wsOrder.cell(rCount,  8).fill = redFill
                wsOrder.cell(rCount,  5).value = 'x'   
                wsOrder.cell(rCount,  6).value = 'x' 
                wsOrder.cell(rCount,  7).value = 'x'


# Desc: Calculates averages of proportions for each grouping (AO Noise, etc.) for famil, pre, center, and post looks
# Params: filePath (str), ordnum (int), wbOrder (Workbook), ID (int), orderFile (str)
# Returns: none
# Side Effects: reads in trial nums; opens and saves the file
def order_calculations(filePath: str, ordnum: int, wbOrder: Workbook, ID: int, orderFile: str) :
    wbOrder.create_sheet("Values Only")
    wbOrder.save(str(filePath + "/Order/" + ID + "_" + orderFile))
    wsRaw = wbOrder.worksheets[0]
    wsValsOnly = wbOrder.worksheets[1]

    with open(filePath + '/Input/Order ' + ordnum + '_Trials.txt') as p:
        rawLines = p.readlines()
        trialTypes = [] # to store column names

        temp = []
        for line in rawLines:
            line = line.strip() # remove leading/trailing white spaces
            if line != '':
                temp.append(line)
            else:
                trialTypes.append(temp)
                temp = []
        trialTypes.append(temp)

    calculatedData = []

    for trials in trialTypes:
        temp = []
        temp.append(trials[0])

        lookCount = 0.0

        i = 1
        numTrials = 0
        while (i < len(trials)):
            look = trials[i][-1]
            if (look == 'L'):
                trialNum = int(trials[i][:-1])
                addThis = wsRaw.cell(trialNum + 4, 5).value
            elif (look == 'R'):
                trialNum = int(trials[i][:-1])
                addThis = wsRaw.cell(trialNum + 4, 6).value
            else:
                trialNum = int(trials[i])
                addThis = wsRaw.cell(trialNum + 4, 7).value
            if (type(addThis) is not str):
                numTrials+=1
                lookCount+=addThis
            i+=1
        
        avgValue = float(lookCount / numTrials)
        temp.append(avgValue)
        calculatedData.append(temp)
    
    colCount = 1
    for data in calculatedData:
        wsValsOnly.cell(1, colCount).value = data[0]
        wsValsOnly.cell(2, colCount).value = data[1]
        colCount+=1

    wbOrder.save(str(filePath + "/Order/" + ID + "_" + orderFile))

    post_to_pre_calculations(filePath, ordnum, wbOrder, ID, orderFile)


# Desc: calculates the averages between postnaming - prenaming looks for each category
# Params: filePath (str), ordNum (int), wbOrder (Workbook), ID (int), orderFile (str)
# Returns: none
# Side Effects: opens post to pre text file; opens and saves files
def post_to_pre_calculations(filePath: str, ordnum: int, wbOrder: Workbook, ID: int, orderFile: str):
    groupings = []
    wsRaw = wbOrder.worksheets[0]
    wsValsOnly = wbOrder.worksheets[1]
    with open(filePath + '/Input/Order ' + ordnum + '_PostToPre.txt') as p:
        rawLines = p.readlines()

        temp = []
        for line in rawLines:
            line = line.strip() # remove leading/trailing white spaces
            if line != '':
                temp.append(line)
            else:
                groupings.append(temp)
                temp = []
        groupings.append(temp)

    calculatedAvgs = []
    
    for trials in groupings:
        calculatedDiffs = []
        mother = []
        mother.append(trials[0])

        i = 1
        while (i < len(trials)):
            pair = trials[i].split()

            temp = []

            for look in pair:
                direction = look[-1]
                trialNum = int(look[:-1])

                if (direction == 'L'):
                    val = wsRaw.cell(trialNum + 4, 5).value
                elif (direction == 'R'):
                    val = wsRaw.cell(trialNum + 4, 6).value
                else:
                    print("Error! Missing look direction in PreToPost file!")
                
                if (type(val) is not str):
                    temp.append(val)
                else:
                    temp.append(0)
            
            diff = temp[0] - temp[1]
            calculatedDiffs.append(diff)
            i+=1

        sum = 0.0
        for diffs in calculatedDiffs:
            sum += diffs
        
        calcAvg = sum / len(calculatedDiffs)
        mother.append(calcAvg)

        calculatedAvgs.append(mother)

    colCount = 15
    for data in calculatedAvgs:
        wsValsOnly.cell(1, colCount).value = data[0]
        wsValsOnly.cell(2, colCount).value = data[1]
        colCount+=1

    wbOrder.save(str(filePath + "/Order/" + ID + "_" + orderFile))


# Desc: moves all order data in 'Values Only' WS into final 'Data' WS
# Params: filePath (str)
# Returns: none
# Side Effects: opens participants; opens and saves files
def extract_order_data(filePath):
    print('Compiling all data...')
    with open(filePath + '/Input/Participants.txt') as p:
        rawLines = p.readlines()
        pcps = [] # To store column names

        i = 1
        temp = []
        for line in rawLines:
            line = line.strip() # remove leading/trailing white spaces
            if line != '':
                temp.append(line)
            else:
                pcps.append(temp)
                temp = []
        pcps.append(temp)

    wbData = openpyxl.Workbook()
    wsData = wbData.active

    create_data_file(wsData)

    rowCount = 3

    for pcp in pcps:
        ID = pcp[0]
        if len(pcp) == 3:
            gender = None
            ordnum = pcp[1]
        else:
            gender = pcp[1]
            age = pcp[2]
            monobili = pcp[3]
            ordnum = pcp[4] 
            medium = pcp[5]

        lang = pcp[len(pcp) - 1]

        orderFile = "Order " + ordnum + "_FaceTalk_" + lang + ".xlsx"

        wbOrder = openpyxl.load_workbook(filePath + "/Order/" + ID + "_" + orderFile)
        wsOrder = wbOrder.worksheets[1]

        data = []

        for col in range(1,wsOrder.max_column+1):
            if(wsOrder.cell(2,col).value is not None):
                data.append(wsOrder.cell(2,col).value)

        colInd = 7
        for d in data:
            wsData.cell(rowCount,  1).value = ID
            if gender:
                wsData.cell(rowCount,  2).value = gender
                wsData.cell(rowCount,  3).value = age
                wsData.cell(rowCount,  4).value = monobili
                wsData.cell(rowCount,  6).value = medium
            wsData.cell(rowCount,  5).value = ordnum
            wsData.cell(rowCount,   colInd).value = d
            colInd = colInd + 1
        rowCount = rowCount + 1

    wbData.save(str(filePath + "/Output/" + "results.xlsx"))


# Desc: creates the headers for the 'Data' worksheet for final product
# Params: wsData (Worksheet)
# Returns: none
# Side Effects: none
def create_data_file(wsData: Worksheet):
    secHeaders = ['Demo Info', 'Familiarization', 'PreNaming', 'Naming (Center Looks)', 'PostNaming', 'Post - Pre naming increase in looking', 'Noise vs NoNoise', 'AV vs AO']
    wsData.cell(1,  1).value = secHeaders[0]
    wsData.cell(1,  7).value = secHeaders[1]
    wsData.cell(1,  9).value = secHeaders[2]
    wsData.cell(1,  13).value = secHeaders[3]
    wsData.cell(1,  17).value = secHeaders[4]
    wsData.cell(1,  21).value = secHeaders[5]
    wsData.cell(1,  25).value = secHeaders[6]
    wsData.cell(1,  27).value = secHeaders[7]

    dataHeaders = ['Participant #', 'Gender', 'Age at Test', 'Mono/Bilingual?', 'Order', 'O/IP', 
        'Left Prop', 'Right Prop', 
        'AO Noise', 'AO NoNoise', 'AV Noise', 'AV NoNoise', 
        'AO Noise', 'AO NoNoise', 'AV Noise', 'AV NoNoise',
        'AO Noise', 'AO NoNoise', 'AV Noise', 'AV NoNoise',
        'AO Noise', 'AO NoNoise', 'AV Noise', 'AV NoNoise',
        'AO', 'AV',
        'Noise', 'NoNoise']

    for i in range(1, len(dataHeaders) + 1):
        wsData.cell(2,  i).value = dataHeaders[i - 1]


# Starting Point
def main():
    print('\nWelcome to Order Analysis!\n')

    print('Please review your files and documentation instructions. Would you like to begin the extracting process now?')
    ans = input('Press 1 if ready.\n')
    if (ans == '1'):
        filePath = get_application_path()
        combined_to_order(filePath)
        extract_order_data(filePath)
    
    print('Goodbye!')

if __name__ == "__main__":
    main()